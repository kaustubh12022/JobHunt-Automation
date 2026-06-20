import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from datetime import datetime
from src.logger import logger
from src.models import Job
from src.config_loader import load_config

def get_human_date_str() -> str:
    day = str(datetime.now().day)
    month = datetime.now().strftime("%B")
    return f"{day} {month}"

def generate_final_sheet(jobs: list[Job], pdf_paths: list[str] = None, test_mode: bool = False):
    if not jobs:
        logger.warning("⚠️ No jobs to write to sheet.")
        return None
        
    config = load_config()
    min_score = 0 if test_mode else config['scoring'].get('minimum_score', 50)
    top_n = 3 if test_mode else config['scoring'].get('top_n', 20)
    
    filtered_jobs = [j for j in jobs if j.score is not None and j.score >= min_score]
    shortlisted = filtered_jobs[:top_n]
    
    logger.info(f"📋 Writing {len(shortlisted)} shortlisted jobs to Excel sheet...")
    
    data = []
    for i, job in enumerate(shortlisted):
        row_data = {
            "Rank": i + 1,
            "Score (%)": job.score,
            "Source": job.source,
            "Company": job.company,
            "Title": job.title,
            "Location": job.location,
            "URL": job.url,
            "Missing Skills": ", ".join(job.missing_skills) if job.missing_skills else "",
            "AI Reasons": job.reasons,
            "Full Description": job.description,
        }
        
        if pdf_paths and i < len(pdf_paths) and pdf_paths[i]:
            pdf_path_obj = Path(pdf_paths[i])
            row_data["Tailored Resume"] = pdf_path_obj.name
        else:
            row_data["Tailored Resume"] = ""
            
        data.append(row_data)
        
    df = pd.DataFrame(data)
    
    today_str = "_test" if test_mode else get_human_date_str()
    out_dir_path = Path(config['output']['desktop_path']) / config['output']['folder_name']
    out_folder = out_dir_path / today_str
    out_folder.mkdir(parents=True, exist_ok=True)
    
    file_path = out_folder / f"AutoApply_Jobs_{today_str}.xlsx"
    df.to_excel(file_path, index=False, engine='openpyxl')
    
    wb = load_workbook(file_path)
    ws = wb.active
    
    header_fill = PatternFill(start_color="2D2D2D", end_color="2D2D2D", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    link_font = Font(color="0563C1", underline="single")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
        
    resume_col_idx = ws.max_column
    url_col_idx = 7
    
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.border = thin_border
            # Allow text wrapping for long content
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            
        score_cell = row[1]
        try:
            val = float(score_cell.value)
            if val >= 80:
                score_cell.fill = green_fill
            elif val >= 50:
                score_cell.fill = yellow_fill
        except:
            pass
            
        resume_cell = row[resume_col_idx - 1]
        filename = resume_cell.value
        if filename and str(filename).strip():
            formula = f'=HYPERLINK("{filename}", "📄 Open Resume")'
            resume_cell.value = formula
            resume_cell.font = link_font
            
        url_cell = row[url_col_idx - 1]
        url_val = url_cell.value
        if url_val and str(url_val).startswith('http'):
            formula = f'=HYPERLINK("{url_val}", "🔗 Apply Now")'
            url_cell.value = formula
            url_cell.font = link_font
            
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = (max_length + 2)
        
        # Descriptions and missing skills get much wider columns to prevent extreme vertical stretching
        if col_letter in ['H', 'I', 'J']: 
            adjusted_width = 80
        elif col_letter in ['E', 'D', 'F']: # Title, Company, Location
            adjusted_width = 45
        elif adjusted_width > 40:
            adjusted_width = 40
            
        ws.column_dimensions[col_letter].width = adjusted_width
        
    # Freeze panes: freeze the header row and first 3 columns
    ws.freeze_panes = "D2"
        
    wb.save(file_path)
    logger.info(f"✅ Final Excel sheet saved: {file_path}")
    return str(file_path)